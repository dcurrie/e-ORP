"""build_taxcalc.py — Generate taxcalc.xlsx workbook.

Requires: pip install openpyxl
Run:      python build_taxcalc.py
Output:   taxcalc.xlsx  (open in Numbers or Excel)
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

# ── Palette ────────────────────────────────────────────────────────────────
C_HEADER_BG  = "2E4057"
C_HEADER_FG  = "FFFFFF"
C_INPUT_BG   = "EAF4FB"
C_CALC_BG    = "F4F4F4"
C_RESULT_BG  = "FFF9E6"
C_LABEL_FG   = "333333"
C_SECTION_BG = "D9E8F5"
C_HELPER_BG  = "E0E0E0"
C_NOTE_BG    = "FFF3CD"
C_WARN_FG    = "7B4F00"

FMT_DOLLAR = '#,##0'
FMT_PCT    = '0.0%'
FMT_INT    = '0'

YEARS = [2025, 2026, 2027]

# ── Bracket data ───────────────────────────────────────────────────────────
# Each entry: (year, [(bracket_top, rate), ...])
# Final row MUST be sentinel: (9_999_999, top_rate).
# All years must have the same number of rows.
# TO ADD A NEW YEAR: append tuple here and re-run script.

ORD_BRACKETS_MFJ = [
    (2025, [(23_850, 0.100), (96_950, 0.120), (206_700, 0.220),
            (394_600, 0.240), (501_050, 0.320), (9_999_999, 0.370)]),
    (2026, [(24_800, 0.100), (100_800, 0.120), (211_400, 0.220),
            (403_550, 0.240), (512_450, 0.320), (9_999_999, 0.370)]),
]
ORD_BRACKETS_SINGLE = [
    (2025, [(11_925, 0.100), (48_475, 0.120), (103_350, 0.220),
            (197_300, 0.240), (250_525, 0.320), (9_999_999, 0.370)]),
    (2026, [(12_400, 0.100), (50_400, 0.120), (105_700, 0.220),
            (201_775, 0.240), (256_225, 0.320), (9_999_999, 0.370)]),
]
CG_BRACKETS_MFJ = [
    (2025, [(96_700, 0.000), (600_050, 0.150), (9_999_999, 0.200)]),
    (2026, [(98_900, 0.000), (613_700, 0.150), (9_999_999, 0.200)]),
]
CG_BRACKETS_SINGLE = [
    (2025, [(48_350, 0.000), (533_400, 0.150), (9_999_999, 0.200)]),
    (2026, [(49_450, 0.000), (545_500, 0.150), (9_999_999, 0.200)]),
]
STD_DEDUCTIONS = [
    # (year, mfj_base, mfj_over65_each, single_base, single_over65_each)
    (2025, 31_500, 1_600, 15_750, 1_600),
    (2026, 32_200, 1_650, 16_100, 1_650),
]

# ── Style helpers ──────────────────────────────────────────────────────────

def _font(bold=False, italic=False, color=C_LABEL_FG, size=11):
    return Font(bold=bold, italic=italic, color=color, size=size)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _hdr(ws, row, col, text, span=1, bg=C_HEADER_BG, fg=C_HEADER_FG, size=11):
    """Write a header row. Text goes in col; remaining span cells get background
    and border only (no merge). This allows the column to be duplicated freely."""
    c = ws.cell(row=row, column=col, value=text)
    c.font      = _font(bold=True, color=fg, size=size)
    c.fill      = _fill(bg)
    c.alignment = _align("left", wrap=False)
    c.border    = _border()
    for extra in range(1, span):
        ec = ws.cell(row=row, column=col + extra, value=None)
        ec.fill   = _fill(bg)
        ec.border = _border()
    return c

def _label(ws, row, col, text, bold=False, italic=False, bg=None, align="left"):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = _font(bold=bold, italic=italic)
    c.alignment = _align(align, wrap=True)
    if bg:
        c.fill = _fill(bg)
    return c

def _note(ws, row, col, text, span=1):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = _font(italic=True, color=C_WARN_FG)
    c.fill      = _fill(C_NOTE_BG)
    c.alignment = _align("left", wrap=True)
    c.border    = _border()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row,   end_column=col + span - 1)
    return c

def _val(ws, row, col, value, fmt=None, bg=C_CALC_BG, editable=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = _fill(C_INPUT_BG if editable else bg)
    c.border    = _border()
    c.alignment = _align("right")
    if fmt:
        c.number_format = fmt
    return c

def _formula(ws, row, col, formula, fmt=None, bg=C_CALC_BG):
    c = ws.cell(row=row, column=col, value=formula)
    c.fill      = _fill(bg)
    c.border    = _border()
    c.alignment = _align("right")
    if fmt:
        c.number_format = fmt
    return c

def _set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ── TaxTables bracket writer ───────────────────────────────────────────────

def write_bracket_table(ws, start_row, col, title, bracket_data):
    n_years = len(bracket_data)
    n_rows  = len(bracket_data[0][1])
    for yr, rows in bracket_data:
        assert len(rows) == n_rows, f"Bracket count mismatch for {yr} in {title}"

    total_cols = 1 + n_years * 2
    _hdr(ws, start_row, col, title, span=total_cols)
    hr = start_row + 1
    _label(ws, hr, col, "Idx", bold=True, bg=C_SECTION_BG)
    for i, (year, _) in enumerate(bracket_data):
        _label(ws, hr, col+1+i*2, f"{year} Top ($)", bold=True, bg=C_SECTION_BG)
        _label(ws, hr, col+2+i*2, f"{year} Rate",    bold=True, bg=C_SECTION_BG)

    data_start = start_row + 2
    for row_i in range(n_rows):
        r = data_start + row_i
        _val(ws, r, col, row_i+1, fmt=FMT_INT)
        for yr_i, (year, rows) in enumerate(bracket_data):
            top, rate = rows[row_i]
            _val(ws, r, col+1+yr_i*2, top,  fmt=FMT_DOLLAR)
            _val(ws, r, col+2+yr_i*2, rate, fmt=FMT_PCT)

    return {
        "n_years":    n_years,
        "n_rows":     n_rows,
        "years":      [y for y, _ in bracket_data],
        "data_start": data_start,
        "data_end":   data_start + n_rows - 1,
        "base_col":   col + 1,
    }


# ── TaxTables global registry ──────────────────────────────────────────────
TT = {}


def build_tax_tables(wb):
    ws = wb.create_sheet("TaxTables")
    ws.sheet_view.showGridLines = True
    last_adj_year = max(y for y, _ in ORD_BRACKETS_MFJ)
    r = 1

    _hdr(ws, r, 1, "TaxTables — Maintenance Guide", span=7, size=13)
    r += 1
    ws.row_dimensions[r].height = 200
    doc = (
        "HOW TO ADD A NEW TAX YEAR\n\n"
        "Each November the IRS publishes inflation-adjusted brackets for the coming year. "
        "When new data is available, update this workbook as follows:\n\n"
        "PREFERRED — via script (build_taxcalc.py):\n"
        "  Append a new (year, [...]) tuple to each of the four bracket lists "
        "(ORD_BRACKETS_MFJ, ORD_BRACKETS_SINGLE, CG_BRACKETS_MFJ, CG_BRACKETS_SINGLE) "
        "and a new row to STD_DEDUCTIONS. Re-run the script. "
        "TaxCalc formulas update automatically.\n\n"
        "MANUAL — spreadsheet only:\n"
        "  1. In each bracket table, add two columns to the right of the last year "
        "     (one for Tops, one for Rates). Copy previous year values and update.\n"
        "  2. In the Standard Deductions table, add a new row for the new year.\n"
        "  3. Update LastAdjYear in Global Parameters to the new year.\n"
        "  4. In TaxCalc, in each helper block row extend the nested-IF to add the "
        "     new year's literal value. Pattern:\n"
        "       =IF(MIN(year,LastAdjYear)<=2025, val_2025,\n"
        "        IF(MIN(year,LastAdjYear)<=2026, val_2026, val_2027))\n"
        "     Repeat for tops, rates, and floors for each bracket level.\n\n"
        "BRACKET STRUCTURE: Each table has one row per bracket level. "
        "The final row is a sentinel (Top = 9,999,999) — do not delete it; "
        "update its Rate if the top marginal rate changes via legislation.\n\n"
        "RATES vs TOPS: Both may change each year. "
        "Tops are adjusted for inflation by the IRS; rates change only via legislation."
    )
    c = ws.cell(row=r, column=1, value=doc)
    c.font      = Font(size=10, color=C_WARN_FG)
    c.fill      = _fill(C_NOTE_BG)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.border    = _border()
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 2

    _hdr(ws, r, 1, "Global Parameters", span=3, size=12)
    r += 1
    params = [
        ("InflationRate",  0.02,          FMT_PCT, "Annual inflation for post-LastAdjYear extrapolation"),
        ("LastAdjYear",    last_adj_year, FMT_INT, "Last year with IRS data — update when new year added"),
        ("OBBBAFirstYear", 2025,          FMT_INT, "First year OBBBA senior SSA deduction applies"),
        ("OBBBALastYear",  2028,          FMT_INT, "Last year OBBBA senior SSA deduction applies"),
    ]
    for name, value, fmt, note in params:
        _label(ws, r, 1, name, bold=True)
        _val(ws, r, 2, value, fmt=fmt, editable=True)
        _label(ws, r, 3, note, italic=True)
        wb.defined_names[name] = DefinedName(name, attr_text=f"TaxTables!$B${r}")
        r += 1
    r += 1

    COL_MFJ = 1
    COL_SNG = 9

    _note(ws, r, COL_MFJ,
          "Ordinary income brackets. Each year: two columns (Top, Rate). "
          "Add two columns per new year — see Maintenance Guide.", span=7)
    r += 1
    TT["ord_mfj"]    = write_bracket_table(ws, r, COL_MFJ, "Ordinary Income — MFJ",    ORD_BRACKETS_MFJ)
    TT["ord_single"] = write_bracket_table(ws, r, COL_SNG, "Ordinary Income — Single", ORD_BRACKETS_SINGLE)
    r += 2 + len(ORD_BRACKETS_MFJ[0][1]) + 2

    _note(ws, r, COL_MFJ,
          "Capital gains brackets. Same structure as ordinary income tables.", span=7)
    r += 1
    TT["cg_mfj"]    = write_bracket_table(ws, r, COL_MFJ, "Capital Gains — MFJ",    CG_BRACKETS_MFJ)
    TT["cg_single"] = write_bracket_table(ws, r, COL_SNG, "Capital Gains — Single", CG_BRACKETS_SINGLE)
    r += 2 + len(CG_BRACKETS_MFJ[0][1]) + 2

    _note(ws, r, COL_MFJ,
          "Standard deductions by year. Add a new row per year; update LastAdjYear.", span=7)
    r += 1
    _hdr(ws, r, 1, "Standard Deductions", span=5)
    r += 1
    for i, h in enumerate(["Year","MFJ Base","MFJ Over-65 Each","Single Base","Single Over-65 Each"]):
        _label(ws, r, 1+i, h, bold=True, bg=C_SECTION_BG)
    r += 1
    TT["std_first_data_row"] = r
    TT["std_n_rows"]         = len(STD_DEDUCTIONS)
    for row_data in STD_DEDUCTIONS:
        for ci, val in enumerate(row_data):
            _val(ws, r, 1+ci, val, fmt=FMT_INT if ci == 0 else FMT_DOLLAR)
        r += 1

    for col, w in [(1,6),(2,12),(3,10),(4,12),(5,10),(6,12),(7,10),
                   (9,6),(10,12),(11,10),(12,12),(13,10),(14,12),(15,10)]:
        _set_col_width(ws, col, w)
    return ws


# ── Reference helpers ──────────────────────────────────────────────────────

def col_range_abs(sheet, row_start, row_end, col):
    c = get_column_letter(col)
    return f"{sheet}!${c}${row_start}:${c}${row_end}"


def scalar_bracket_formula(row_i, col_offset, year_cell_ref, fs_cell_ref,
                            bd_mfj, bd_sng, ifact_ref=None):
    """Scalar nested-IF formula selecting one bracket value by year and FS.

    col_offset: 0 = top, 1 = rate  (index into each bracket tuple).
    ifact_ref: if provided, multiplies tops by the inflation factor
               (rates are never inflated — only tops change with inflation).
    All values are embedded as literals — no cross-sheet references.
    """
    val_idx = col_offset
    years   = [y for y, _ in bd_mfj]

    def yr_nested_if(bd):
        vals = [rows[row_i][val_idx] for _, rows in bd]
        if len(vals) == 1:
            return str(vals[0])
        expr = str(vals[-1])
        for i in range(len(years) - 2, -1, -1):
            expr = (f"IF(MIN({year_cell_ref},LastAdjYear)<={years[i]},"
                    f"{vals[i]},{expr})")
        return expr

    base = f"=IF({fs_cell_ref}=1,{yr_nested_if(bd_mfj)},{yr_nested_if(bd_sng)})"
    # Apply inflation factor to tops (col_offset=0) but not rates (col_offset=1)
    if ifact_ref and col_offset == 0:
        base = base + f"*{ifact_ref}"
    return base


def scalar_floor_formula(row_i, year_cell_ref, fs_cell_ref,
                          bd_mfj, bd_sng, ifact_ref=None):
    """Scalar nested-IF selecting the floor for bracket row_i by year and FS.
    Floor = top of previous bracket, 0 for first bracket.
    ifact_ref: if provided, multiplies non-zero floors by inflation factor.
    All values embedded as literals.
    """
    years = [y for y, _ in bd_mfj]

    def floor_val(bd, yr):
        tops = [x[0] for x in dict(bd)[yr]]
        return ([0] + tops[:-1])[row_i]

    def yr_nested_if(bd):
        vals = [floor_val(bd, yr) for yr in years]
        if len(vals) == 1:
            return str(vals[0])
        expr = str(vals[-1])
        for i in range(len(years) - 2, -1, -1):
            expr = (f"IF(MIN({year_cell_ref},LastAdjYear)<={years[i]},"
                    f"{vals[i]},{expr})")
        return expr

    base = f"=IF({fs_cell_ref}=1,{yr_nested_if(bd_mfj)},{yr_nested_if(bd_sng)})"
    # First bracket floor is always 0 — no inflation needed
    if ifact_ref and row_i > 0:
        base = base + f"*{ifact_ref}"
    return base


# ── TaxCalc sheet ──────────────────────────────────────────────────────────

def build_taxcalc(wb):
    ws = wb.create_sheet("TaxCalc")
    ws.sheet_view.showGridLines = True
    SCEN_COL = 2   # column B = scenario 1

    def b(row):
        """Row-absolute, column-relative — follows column when copied."""
        return f"{get_column_letter(SCEN_COL)}${row}"

    r = 1

    # ── Input block ────────────────────────────────────────────
    _hdr(ws, r, 1, "Inputs", span=2, size=12)
    _hdr(ws, r, 3, "Notes", bg=C_SECTION_BG, fg=C_LABEL_FG)
    r += 1

    input_defs = [
        ("Tax Year",                         2025, FMT_INT,    "Enter tax year"),
        ("Filing Status  (0=Single  1=MFJ)", 1,    FMT_INT,    "0 = Single,  1 = Married Filing Jointly"),
        ("Net Earnings",                     0,    FMT_DOLLAR, "Net pay, excluding pre-tax deductions e.g. 401k"),
        ("IRA Distribution",                 0,    FMT_DOLLAR, "Taxable IRA distributions, including QCD amount"),
        ("QCD",                              0,    FMT_DOLLAR, "Qualified Charitable Distribution"),
        ("Interest",                         0,    FMT_DOLLAR, "Taxable interest income"),
        ("Dividends",                        0,    FMT_DOLLAR, "Dividend income"),
        ("SSA Income",                       0,    FMT_DOLLAR, "Gross Social Security benefit"),
        ("ST Capital Gains",                 0,    FMT_DOLLAR, "Short-term capital gains"),
        ("LT Capital Gains",                 0,    FMT_DOLLAR, "Long-term capital gains"),
        ("Gains Carryover",                  0,    FMT_DOLLAR, "LT cap gains carryover from prior year (may be negative)"),
        ("Inflation Rate",                   None, FMT_PCT,    "Linked from TaxTables — do not edit here"),
    ]

    R = {}
    for i, (label, default, fmt, note) in enumerate(input_defs):
        row = r + i
        R[i] = row
        _label(ws, row, 1, label)
        if default is None:
            _formula(ws, row, SCEN_COL, "=InflationRate", fmt=fmt)
        else:
            _val(ws, row, SCEN_COL, default, fmt=fmt, editable=True)
        _label(ws, row, 3, note, italic=True)

    R_YEAR = R[0];  R_FS   = R[1];  R_NPAY = R[2]
    R_IRAD = R[3];  R_QCD  = R[4];  R_INTR = R[5]
    R_DVDD = R[6];  R_SSAI = R[7];  R_STGN = R[8]
    R_LTGN = R[9];  R_GCRY = R[10]; R_INFL = R[11]
    r += len(input_defs) + 1

    # ── Intermediate calculations ──────────────────────────────
    _hdr(ws, r, 1, "Intermediate Calculations", span=2, size=12)
    r += 1

    def calc_row(label, formula, fmt, note="", bg=C_CALC_BG):
        nonlocal r
        _label(ws, r, 1, label)
        _formula(ws, r, SCEN_COL, formula, fmt=fmt, bg=bg)
        if note:
            _label(ws, r, 3, note, italic=True)
        row_ref = r; r += 1
        return row_ref

    R_TXSSA = calc_row("Taxable SSA (85%)",
                       f"={b(R_SSAI)}*0.85", FMT_DOLLAR,
                       "85% of SSA income is taxable")
    R_INC   = calc_row("Income (ex cap gains)",
                       f"={b(R_NPAY)}+{b(R_IRAD)}-{b(R_QCD)}+{b(R_INTR)}+{b(R_DVDD)}+{b(R_TXSSA)}+{b(R_STGN)}",
                       FMT_DOLLAR)
    R_CG    = calc_row("Cap Gains (net)",
                       f"={b(R_LTGN)}+{b(R_GCRY)}", FMT_DOLLAR,
                       "Net of carryover; may be negative")
    R_AGI   = calc_row("AGI",   f"={b(R_INC)}+{b(R_CG)}", FMT_DOLLAR)
    R_MAGI  = calc_row("MAGI",
                       f"={b(R_AGI)}+{b(R_SSAI)}*0.15+{b(R_QCD)}", FMT_DOLLAR,
                       "Adds non-taxable SSA portion and QCD back")
    R_PAX   = calc_row("OBBBA_pax",
                       f"=IF({b(R_FS)}=1,2,1)", FMT_INT,
                       "Qualifying seniors: 1 = Single, 2 = MFJ")
    R_EXC   = calc_row("OBBBA_excess",
                       f"=MAX(0,{b(R_MAGI)}-{b(R_PAX)}*75000)", FMT_DOLLAR,
                       "MAGI above phase-out threshold ($75k per person)")
    R_OBDED = calc_row("OBBBA_deduction",
                       f"=IF(AND({b(R_YEAR)}>=OBBBAFirstYear,{b(R_YEAR)}<=OBBBALastYear),"
                       f"MAX(0,{b(R_PAX)}*6000-{b(R_PAX)}*0.06*{b(R_EXC)}),0)",
                       FMT_DOLLAR,
                       "Senior SSA deduction 2025–2028; phases out above threshold")
    R_IFACT = calc_row("Inflation Factor",
                       f"=(1+{b(R_INFL)})^MAX(0,{b(R_YEAR)}-LastAdjYear)",
                       "0.000",
                       "1.0 for years with IRS data; compounds forward beyond LastAdjYear")

    std_r1  = TT["std_first_data_row"]
    std_r2  = std_r1 + TT["std_n_rows"] - 1
    yr_col  = col_range_abs("TaxTables", std_r1, std_r2, 1)
    mfj_b   = col_range_abs("TaxTables", std_r1, std_r2, 2)
    mfj_o65 = col_range_abs("TaxTables", std_r1, std_r2, 3)
    sng_b   = col_range_abs("TaxTables", std_r1, std_r2, 4)
    sng_o65 = col_range_abs("TaxTables", std_r1, std_r2, 5)
    lkp_yr  = f"MIN({b(R_YEAR)},LastAdjYear)"

    R_SDBASE = calc_row("Std Deduction (base)",
                        f"=IF({b(R_FS)}=1,"
                        f"XLOOKUP({lkp_yr},{yr_col},{mfj_b}),"
                        f"XLOOKUP({lkp_yr},{yr_col},{sng_b}))*{b(R_IFACT)}",
                        FMT_DOLLAR, "Base deduction × inflation factor")
    R_SDO65  = calc_row("Std Deduction (over-65)",
                        f"={b(R_PAX)}*IF({b(R_FS)}=1,"
                        f"XLOOKUP({lkp_yr},{yr_col},{mfj_o65}),"
                        f"XLOOKUP({lkp_yr},{yr_col},{sng_o65}))*{b(R_IFACT)}",
                        FMT_DOLLAR, "Over-65 addition × pax × inflation factor")
    R_SD     = calc_row("Std Deduction (total)",
                        f"={b(R_SDBASE)}+{b(R_SDO65)}+{b(R_OBDED)}",
                        FMT_DOLLAR, "Base + over-65 + OBBBA deduction")

    R_TXINC = calc_row("Taxable Income (ordinary)",
                       f"=MAX(0,{b(R_INC)}-{b(R_SD)})",
                       FMT_DOLLAR,
                       "Income after std deduction; distinct from AGI")
    r += 1

    # ── Per-bracket helper blocks ──────────────────────────────
    # Each block has one row per bracket level.
    # Every cell is a plain scalar — no arrays, no range operations.
    # Numbers does not support implicit array expansion of scalar-range
    # operations inside SUMPRODUCT when imported from xlsx, so we avoid
    # all range arithmetic entirely.
    #
    # Block order per table:  top | rate | floor | ord_tax | cg_tax
    # Tax blocks reference their own table's top/rate/floor rows by
    # row-absolute column-relative addresses, so they copy with the column.

    _hdr(ws, r, 1, "Per-Bracket Helper Blocks — copy with scenario column",
         span=3, bg=C_HEADER_BG, fg=C_HEADER_FG)
    r += 1
    _label(ws, r, 1,
           "One row per bracket level. Every cell is a plain scalar formula "
           "(no arrays). Tax rows sum to give total tax on income / cap gains.",
           italic=True)
    r += 1

    year_ref = b(R_YEAR)
    fs_ref   = b(R_FS)

    def write_scalar_block(label, fmt, formula_fn, n):
        """Write n scalar helper rows; return start row."""
        nonlocal r
        start = r
        for i in range(n):
            _label(ws, r, 1, f"{label} {i+1}", italic=True)
            _formula(ws, r, SCEN_COL, formula_fn(i), fmt=fmt, bg=C_HELPER_BG)
            r += 1
        return start

    # ── Ordinary income blocks ─────────────────────────────────
    n_ord = len(ORD_BRACKETS_MFJ[0][1])

    ifact_ref = b(R_IFACT)

    R_ORD_TOP_START = write_scalar_block(
        "Ord top", FMT_DOLLAR, n=n_ord,
        formula_fn=lambda i: scalar_bracket_formula(
            i, 0, year_ref, fs_ref, ORD_BRACKETS_MFJ, ORD_BRACKETS_SINGLE,
            ifact_ref=ifact_ref))

    R_ORD_RATE_START = write_scalar_block(
        "Ord rate", FMT_PCT, n=n_ord,
        formula_fn=lambda i: scalar_bracket_formula(
            i, 1, year_ref, fs_ref, ORD_BRACKETS_MFJ, ORD_BRACKETS_SINGLE))

    R_ORD_FLOOR_START = write_scalar_block(
        "Ord floor", FMT_DOLLAR, n=n_ord,
        formula_fn=lambda i: scalar_floor_formula(
            i, year_ref, fs_ref, ORD_BRACKETS_MFJ, ORD_BRACKETS_SINGLE,
            ifact_ref=ifact_ref))

    # Ordinary tax per bracket:
    # MAX(0, MIN(taxable_income, top_i) - floor_i) * rate_i
    # All three operands are scalar cell refs — no range arithmetic.
    def ord_tax_formula(i):
        top   = b(R_ORD_TOP_START   + i)
        floor = b(R_ORD_FLOOR_START + i)
        rate  = b(R_ORD_RATE_START  + i)
        inc   = b(R_TXINC)
        return f"=MAX(0,MIN({inc},{top})-{floor})*{rate}"

    R_ORD_TAX_START = write_scalar_block(
        "Ord tax bracket", FMT_DOLLAR, n=n_ord,
        formula_fn=ord_tax_formula)

    r += 1  # spacer

    # ── Capital gains blocks ───────────────────────────────────
    # Cap gains are stacked on top of ordinary taxable income for rate
    # determination: the base for CG brackets is R_TXINC, and income
    # within each CG bracket = MAX(0, MIN(txinc+cg, top_i) - floor_i)
    #                         - MAX(0, MIN(txinc,    top_i) - floor_i)
    # This gives only the portion attributable to cap gains in each bracket.

    n_cg = len(CG_BRACKETS_MFJ[0][1])

    R_CG_TOP_START = write_scalar_block(
        "CG top", FMT_DOLLAR, n=n_cg,
        formula_fn=lambda i: scalar_bracket_formula(
            i, 0, year_ref, fs_ref, CG_BRACKETS_MFJ, CG_BRACKETS_SINGLE,
            ifact_ref=ifact_ref))

    R_CG_RATE_START = write_scalar_block(
        "CG rate", FMT_PCT, n=n_cg,
        formula_fn=lambda i: scalar_bracket_formula(
            i, 1, year_ref, fs_ref, CG_BRACKETS_MFJ, CG_BRACKETS_SINGLE))

    R_CG_FLOOR_START = write_scalar_block(
        "CG floor", FMT_DOLLAR, n=n_cg,
        formula_fn=lambda i: scalar_floor_formula(
            i, year_ref, fs_ref, CG_BRACKETS_MFJ, CG_BRACKETS_SINGLE,
            ifact_ref=ifact_ref))

    def cg_tax_formula(i):
        top    = b(R_CG_TOP_START   + i)
        floor  = b(R_CG_FLOOR_START + i)
        rate   = b(R_CG_RATE_START  + i)
        base   = b(R_TXINC)
        cg_net = f"MAX(0,{b(R_CG)})"
        total  = f"({base}+{cg_net})"
        # tax on (ordinary+cg) in this bracket minus tax on ordinary alone
        return (f"=(MAX(0,MIN({total},{top})-{floor})"
                f"-MAX(0,MIN({base},{top})-{floor}))*{rate}")

    R_CG_TAX_START = write_scalar_block(
        "CG tax bracket", FMT_DOLLAR, n=n_cg,
        formula_fn=cg_tax_formula)

    r += 1  # spacer

    # ── Tax calculation rows ───────────────────────────────────
    _hdr(ws, r, 1, "Tax Calculation", span=2, size=12)
    r += 1

    def sum_block(start_row, n):
        """Column-relative SUM over n contiguous rows from start_row."""
        col = get_column_letter(SCEN_COL)
        return f"SUM({col}${start_row}:{col}${start_row + n - 1})"

    R_IBTAX = calc_row("Tax on Income",
                       f"={sum_block(R_ORD_TAX_START, n_ord)}",
                       FMT_DOLLAR,
                       "Sum of per-bracket ordinary income tax")

    # Marginal rate: find last bracket where floor < taxable income
    def marginal_rate_formula(txinc_ref, top_start, rate_start, n):
        # Walk brackets from last to first; return rate of highest bracket reached
        # Built as nested IFs: IF(txinc>floor_n, rate_n, IF(txinc>floor_{n-1}, ...))
        # Floors are already in helper rows; use them directly.
        expr = "0"
        for i in range(n - 1, -1, -1):
            floor = b(R_ORD_FLOOR_START + i)
            rate  = b(rate_start + i)
            expr  = f"IF({txinc_ref}>{floor},{rate},{expr})"
        return f"={expr}"

    # Marginal rate: nest highest floor first; first true branch wins.
    # pairs = [(floor_ref, rate_ref)] for brackets 1..n-1, highest last.
    # Wrap reversed so outermost IF checks highest floor.
    def mrate_formula(txinc_ref, floor_start, rate_start, n):
        # innermost default = rate for bracket 0 (floor=0, always reached)
        # wrap outward from bracket 1 to n-1, highest bracket outermost
        # so outermost IF checks highest floor first
        expr = b(rate_start)                    # bracket 0: innermost default
        for i in range(1, n):                   # brackets 1..n-1, low to high
            floor = b(floor_start + i)
            rate  = b(rate_start  + i)
            expr  = f"IF({txinc_ref}<={floor},{expr},{rate})"  # true=lower, false=higher
        return f"=IF({txinc_ref}>0,{expr},0)"

    def cg_mrate_formula(n):
        cg_net = f"MAX(0,{b(R_CG)})"
        total  = f"({b(R_TXINC)}+{cg_net})"
        expr   = b(R_CG_RATE_START)
        for i in range(1, n):
            floor = b(R_CG_FLOOR_START + i)
            rate  = b(R_CG_RATE_START  + i)
            expr  = f"IF({total}<={floor},{expr},{rate})"
        return f"=IF({cg_net}=0,0,IF({total}>0,{expr},0))"

    R_MRATE = calc_row("Marginal Rate — Income",
                       mrate_formula(b(R_TXINC),
                                     R_ORD_FLOOR_START,
                                     R_ORD_RATE_START, n_ord),
                       FMT_PCT)

    R_BREND = calc_row("Income Bracket Ceiling",
                       # Top of the bracket containing taxable income
                       "=" + "".join(
                           f"IF({b(R_TXINC)}<={b(R_ORD_TOP_START+i)},{b(R_ORD_TOP_START+i)},"
                           for i in range(n_ord)
                       ) + "0" + ")" * n_ord,
                       FMT_DOLLAR)

    R_CGTAX = calc_row("Tax on Cap Gains",
                       f"={sum_block(R_CG_TAX_START, n_cg)}",
                       FMT_DOLLAR,
                       "Sum of per-bracket cap gains tax; only positive net CG taxed")



    R_CRATE = calc_row("Marginal Rate — Cap Gains",
                       cg_mrate_formula(n_cg),
                       FMT_PCT)
    r += 1

    # ── Results summary ────────────────────────────────────────
    _hdr(ws, r, 1, "Results Summary", span=2, size=12,
         bg="1D6A96", fg="FFFFFF")
    r += 1

    def result_row(label, formula, fmt, note=""):
        nonlocal r
        _label(ws, r, 1, label, bold=True)
        _formula(ws, r, SCEN_COL, formula, fmt=fmt, bg=C_RESULT_BG)
        if note:
            _label(ws, r, 3, note, italic=True)
        row_ref = r; r += 1
        return row_ref

    global TC_RESULTS
    TC_RESULTS = {}
    TC_RESULTS["Income"]     = result_row("Income (ex cap gains)", f"=ROUND({b(R_INC)},0)",   FMT_DOLLAR)
    TC_RESULTS["Cap Gains"]  = result_row("Cap Gains",             f"=ROUND({b(R_CG)},0)",    FMT_DOLLAR)
    TC_RESULTS["AGI"]        = result_row("AGI",                   f"=ROUND({b(R_AGI)},0)",   FMT_DOLLAR)
    TC_RESULTS["Std Deduct"] = result_row("Standard Deduction",    f"=ROUND({b(R_SD)},0)",    FMT_DOLLAR)
    TC_RESULTS["Total Tax"]  = result_row("Total Tax",
                                          f"=ROUND({b(R_IBTAX)}+{b(R_CGTAX)},0)",            FMT_DOLLAR,
                                          "Rounded to nearest dollar")
    TC_RESULTS["Tax Income"] = result_row("Tax on Income",         f"=ROUND({b(R_IBTAX)},0)", FMT_DOLLAR)
    TC_RESULTS["Tax CG"]     = result_row("Tax on Cap Gains",      f"=ROUND({b(R_CGTAX)},0)", FMT_DOLLAR)
    TC_RESULTS["M Rate Inc"] = result_row("Marginal Rate — Income",    f"={b(R_MRATE)}",      FMT_PCT)
    TC_RESULTS["M Rate CG"]  = result_row("Marginal Rate — Cap Gains", f"={b(R_CRATE)}",      FMT_PCT)
    TC_RESULTS["Eff Rate"]   = result_row("Effective Rate",
                                          f"=IFERROR(ROUND({b(TC_RESULTS['Total Tax'])}"
                                          f"/{b(R_AGI)},4),0)",                               FMT_PCT)

    _set_col_width(ws, 1, 32)
    _set_col_width(ws, 2, 18)
    _set_col_width(ws, 3, 56)
    return ws


# ── Year sheets ────────────────────────────────────────────────────────────

def build_year_sheet(wb, year):
    ws = wb.create_sheet(str(year))
    _hdr(ws, 1, 1, f"Tax Year {year}", span=3, size=13)
    r = 3
    _hdr(ws, r, 1, "Inputs — Scenario 1", span=2, bg=C_SECTION_BG, fg=C_LABEL_FG)
    _label(ws, r, 3, "Link these cells to TaxCalc column B rows 2–12", italic=True)
    r += 1

    for label, default, fmt, note in [
        ("Tax Year",         year, FMT_INT,    ""),
        ("Filing Status",    1,    FMT_INT,    "0 = Single   1 = MFJ"),
        ("Net Earnings",     0,    FMT_DOLLAR, ""),
        ("IRA Distribution", 0,    FMT_DOLLAR, ""),
        ("QCD",              0,    FMT_DOLLAR, ""),
        ("Interest",         0,    FMT_DOLLAR, ""),
        ("Dividends",        0,    FMT_DOLLAR, ""),
        ("SSA Income",       0,    FMT_DOLLAR, ""),
        ("ST Capital Gains", 0,    FMT_DOLLAR, ""),
        ("LT Capital Gains", 0,    FMT_DOLLAR, ""),
        ("Gains Carryover",  0,    FMT_DOLLAR, "May be negative"),
    ]:
        _label(ws, r, 1, label)
        _val(ws, r, 2, default, fmt=fmt, editable=True)
        if note:
            _label(ws, r, 3, note, italic=True)
        r += 1

    r += 1
    _hdr(ws, r, 1, "Results — Scenario 1", span=2, bg="1D6A96", fg="FFFFFF")
    _label(ws, r, 3, "Link these cells to TaxCalc Results Summary column B", italic=True)
    r += 1

    for label, fmt in [
        ("Income (ex cap gains)",     FMT_DOLLAR),
        ("Cap Gains",                 FMT_DOLLAR),
        ("AGI",                       FMT_DOLLAR),
        ("Standard Deduction",        FMT_DOLLAR),
        ("Total Tax",                 FMT_DOLLAR),
        ("Tax on Income",             FMT_DOLLAR),
        ("Tax on Cap Gains",          FMT_DOLLAR),
        ("Marginal Rate — Income",    FMT_PCT),
        ("Marginal Rate — Cap Gains", FMT_PCT),
        ("Effective Rate",            FMT_PCT),
    ]:
        _label(ws, r, 1, label, bold=True)
        _val(ws, r, 2, 0, fmt=fmt)
        r += 1

    _set_col_width(ws, 1, 28)
    _set_col_width(ws, 2, 16)
    _set_col_width(ws, 3, 40)
    return ws


# ── Main ───────────────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    wb.remove(wb.active)
    build_tax_tables(wb)
    build_taxcalc(wb)
    for year in YEARS:
        build_year_sheet(wb, year)
    out = "taxcalc.xlsx"
    wb.save(out)
    print(f"Saved {out}")
    print("Open in Numbers, then for each year sheet:")
    print("  1. Link input cells to TaxCalc column B rows 2–12")
    print("  2. Link result cells to TaxCalc Results Summary rows")
    print("  3. Run validation scenarios from the plan")


if __name__ == "__main__":
    main()
